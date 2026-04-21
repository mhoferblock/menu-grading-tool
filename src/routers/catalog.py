from __future__ import annotations

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from src.config import settings
from src.utils.logging import get_logger

log = get_logger("catalog")

router = APIRouter(prefix="/catalog", tags=["catalog"])


@router.post("/fetch")
async def fetch_square_catalog(merchant_id: str = Form(...), market: str = Form("US")):
    """Fetch catalog items from Square API for a given merchant."""
    if not settings.SQUARE_ACCESS_TOKEN:
        log.warning("square_token_missing", merchant_id=merchant_id)
        raise HTTPException(
            status_code=503,
            detail="Square API is not configured. Add SQUARE_ACCESS_TOKEN to enable catalog fetches.",
        )

    try:
        import httpx

        headers = {
            "Square-Version": "2024-09-19",
            "Authorization": f"Bearer {settings.SQUARE_ACCESS_TOKEN}",
            "Content-Type": "application/json",
        }

        base_url = "https://connect.squareup.com/v2"
        items = []
        cursor = None

        async with httpx.AsyncClient(timeout=30) as client:
            for _ in range(10):
                body: dict = {"types": ["ITEM"]}
                if cursor:
                    body["cursor"] = cursor

                resp = await client.post(
                    f"{base_url}/catalog/list",
                    headers=headers,
                    json=body,
                )

                if resp.status_code == 401:
                    raise HTTPException(
                        status_code=401,
                        detail="Square API authentication failed. Check your access token.",
                    )
                if resp.status_code != 200:
                    raise HTTPException(
                        status_code=502,
                        detail=f"Square API error: {resp.status_code} — {resp.text[:200]}",
                    )

                data = resp.json()
                objects = data.get("objects", [])

                for obj in objects:
                    item_data = obj.get("item_data", {})
                    variations = item_data.get("variations", [])
                    price = None
                    if variations:
                        price_money = (
                            variations[0]
                            .get("item_variation_data", {})
                            .get("price_money")
                        )
                        if price_money:
                            price = price_money.get("amount", 0) / 100

                    items.append({
                        "id": obj.get("id"),
                        "name": item_data.get("name", "Unknown"),
                        "description": item_data.get("description", ""),
                        "category_id": item_data.get("category_id"),
                        "price": price,
                        "variation_count": len(variations),
                    })

                cursor = data.get("cursor")
                if not cursor:
                    break

        log.info("catalog_fetched", merchant_id=merchant_id, item_count=len(items))

        return {
            "data": {
                "merchant_id": merchant_id,
                "market": market,
                "item_count": len(items),
                "items": items,
            },
            "meta": {"source": "square_api"},
        }

    except HTTPException:
        raise
    except Exception as e:
        log.error("catalog_fetch_error", merchant_id=merchant_id, error=str(e))
        raise HTTPException(status_code=500, detail=f"Failed to fetch catalog: {str(e)}")


@router.post("/upload")
async def upload_catalog_excel(file: UploadFile = File(...)):
    """Parse an uploaded Excel/CSV catalog file."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{ext}. Use .xlsx, .xls, or .csv")

    contents = await file.read()
    items = []

    try:
        if ext == "csv":
            import csv
            import io
            text = contents.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                name = row.get("name") or row.get("Name") or row.get("Item Name") or row.get("item_name") or ""
                price_str = row.get("price") or row.get("Price") or row.get("amount") or ""
                price = None
                if price_str:
                    try:
                        price = float(str(price_str).replace("$", "").replace(",", "").strip())
                    except ValueError:
                        pass
                category = row.get("category") or row.get("Category") or row.get("category_name") or ""
                description = row.get("description") or row.get("Description") or ""

                if name.strip():
                    items.append({
                        "name": name.strip(),
                        "description": description.strip(),
                        "category": category.strip(),
                        "price": price,
                    })
        else:
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Excel file has no active worksheet")

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise HTTPException(status_code=400, detail="Excel file has no data rows")

            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            name_col = next((i for i, h in enumerate(headers) if h in ("name", "item name", "item_name")), None)
            price_col = next((i for i, h in enumerate(headers) if h in ("price", "amount")), None)
            cat_col = next((i for i, h in enumerate(headers) if h in ("category", "category_name", "category name")), None)
            desc_col = next((i for i, h in enumerate(headers) if h in ("description", "desc")), None)

            if name_col is None:
                name_col = 0

            for row in rows[1:]:
                name = str(row[name_col]).strip() if name_col is not None and row[name_col] else ""
                price = None
                if price_col is not None and row[price_col]:
                    try:
                        price = float(str(row[price_col]).replace("$", "").replace(",", "").strip())
                    except ValueError:
                        pass
                category = str(row[cat_col]).strip() if cat_col is not None and row[cat_col] else ""
                description = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else ""

                if name and name.lower() != "none":
                    items.append({
                        "name": name,
                        "description": description,
                        "category": category,
                        "price": price,
                    })
            wb.close()

    except HTTPException:
        raise
    except Exception as e:
        log.error("catalog_parse_error", filename=file.filename, error=str(e))
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    log.info("catalog_uploaded", filename=file.filename, item_count=len(items))

    return {
        "data": {
            "filename": file.filename,
            "item_count": len(items),
            "items": items,
        },
        "meta": {"source": "excel_upload"},
    }
